import os
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.firefox.service import Service

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from helpers.setup_scraper import setup_scraper
from helpers.scrape_player_data import scrape_player_data
from helpers.get_criteria_from_user import get_criteria_from_user
from helpers.should_highlight_player import should_highlight_player
from helpers.get_rating_name import get_rating_name
from helpers.highlight_row_red import highlight_row_red

load_dotenv()

# ОСНОВНОЙ КОД ПРОГРАММЫ - Я ХЗ КАК MAIN В PYTHON сделать 

filename =  os.environ.get('PATH_TO_CSV')
df_players = pd.read_csv(filename)
players_cfn = df_players["CFN ЧИСЛОВОЙ Код игрока"].to_list()
bad_players = []
if players_cfn:
    print("CFN успешно загружены:")
    print(players_cfn)
else:
    print("Список CFN пуст или произошла ошибка.")

min_matches, max_matches, max_rating = get_criteria_from_user()


# Set up the path to geckodriver if it's not in your system's PATH
geckodriver_path = os.environ.get('GECKODRIVER_PATH')  # Replace with the actual OS path
service = Service(geckodriver_path)

username = os.environ.get('CAPCOM_LOGIN')
password = os.environ.get('CAPCOM_PASSWORD')
# Initialize the Firefox WebDriver
driver = webdriver.Firefox(service=service)
setup_scraper(driver, username, password)
    
# Дальнейшая работа с players_cfn
players_data = []
for cfn in players_cfn:
    player_data = scrape_player_data(driver, cfn, df_players, bad_players)
    if player_data != None:
        players_data.append(player_data)
    #time.sleep(10000)

#Завершаем работу - для дебага можно писать большой time.Sleep(), чтобы походить и взять xpath и прочее
driver.quit()

df = pd.DataFrame(bad_players, columns=["Wrong CFN", "Bad player Nickname"])
df.to_csv('bad_players.csv', index=False)

#EXCEL
#Подготовка к работе с excel - здесь gpt помог сильно 
data = []
for player in players_data:
    for phase, character_data in player.phases.items():
        for i, (char_name, lp) in enumerate(character_data):
            rating_name = get_rating_name(lp)  # Get the rating name based on the LP
            data.append({
                'Player Name': player.name if i == 0 else "",  # Merge player name for the first row
                'CFN': player.cfn if i == 0 else "",  # Merge CFN for the first row
                'Phase': phase if i == 0 else "",  # Merge phase for the first row
                'League Point Name': char_name,
                'League Point LP': lp,
                'Rating Name': rating_name,  # Add the rating name column
                'Ranked Matches': player.matches[0],
                'Casual Matches': player.matches[1],
                'Room Matches': player.matches[2],
                'Battlehub Matches': player.matches[3],
                'Ranked Time': player.played_time[0],
                'Practice Time': player.played_time[1],
                'Custom Room Time': player.played_time[2]
            })

# Дальше gpt - я не прокоментирую сильно - есть комменты на английском

# Create the DataFrame from the data list
df = pd.DataFrame(data)

# Define the Excel filename
excel_filename = "players_league_points_phases_with_rating.xlsx"

# Save the DataFrame to an Excel file
df.to_excel(excel_filename, index=False)

# Now, let's merge cells for each player's CFN and Player Name in the Excel file
workbook = load_workbook(excel_filename)
worksheet = workbook.active

num_columns = len(df.columns)

# Keep track of the starting row for each player and merge cells
current_row = 2  # Since row 1 has headers, start from the second row
for player in players_data:
    total_matches = sum(player.matches)

    # Check if the player should be highlighted
    max_lp = player.max_lp
    highlight = should_highlight_player(total_matches, max_lp, min_matches, max_matches, max_rating)
    challonge_nickname = df_players.loc[df_players["CFN ЧИСЛОВОЙ Код игрока"] == player.cfn, "Participant Username"].values[0]
    if challonge_nickname.lower() != player.name.lower():
        highlight = True

    num_phases = sum(len(character_data) for phase, character_data in player.phases.items())

    if highlight:
            for row in range(current_row, current_row + num_phases):
                highlight_row_red(worksheet, row, num_columns)

    # Merge Player Name cells (Column A)
    worksheet.merge_cells(f'A{current_row}:A{current_row + num_phases - 1}')
    # Merge CFN cells (Column B)
    worksheet.merge_cells(f'B{current_row}:B{current_row + num_phases - 1}')

    worksheet.merge_cells(f'G{current_row}:G{current_row + num_phases - 1}')
    worksheet.merge_cells(f'H{current_row}:H{current_row + num_phases - 1}')
    worksheet.merge_cells(f'I{current_row}:I{current_row + num_phases - 1}')
    worksheet.merge_cells(f'J{current_row}:J{current_row + num_phases - 1}')

    worksheet.merge_cells(f'K{current_row}:K{current_row + num_phases - 1}')
    worksheet.merge_cells(f'L{current_row}:L{current_row + num_phases - 1}')
    worksheet.merge_cells(f'M{current_row}:M{current_row + num_phases - 1}')
    
    # Move to the next player's data
    current_row += num_phases

# Set column widths to auto-adjust based on maximum content length
for col in worksheet.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)  # Get the column letter
    for cell in col:
        try:
            # Measure the length of each cell value
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    # Adjust column width based on the maximum content length, with a buffer
    adjusted_width = (max_length + 2)  # Add some extra space for readability
    worksheet.column_dimensions[col_letter].width = adjusted_width

# Save the workbook after merging
workbook.save(excel_filename)

print(f"Data saved to {excel_filename} with merged cells for Player Names, CFNs, and Rating Names.")
