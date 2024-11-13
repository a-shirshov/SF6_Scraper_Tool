from openpyxl.styles import PatternFill

def highlight_row_red(worksheet, row_number, num_columns):
    light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red color
    worksheet.cell(row=row_number, column=1).fill = light_red_fill  # Column 1 corresponds to 'Player Name'
